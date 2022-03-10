from tkinter import *

import openpyxl
from openpyxl.chart import (PieChart, Reference)
import win32com.client as client

root = Tk()
root.geometry("600x400")
root.resizable(0, 0)
root.title("Test Cases Parser")
Label(root, text="Test Cases Parser", font="arial 15 bold").pack()

tester_label = Label(root, text="Tester Name", font="arial 10 bold").pack()
tester_str = StringVar()
Entry(root, textvariable=tester_str).pack()

path = r"C:\Users\danie\OneDrive\Desktop\Proiect\Test_Case_Format_afisor date.xlsx"
path_PDF = r"C:\Users\danie\OneDrive\Desktop\Proiect\Test_Case_Format_afisor date.pdf"
values = [0, 0]


def compareValues():
    wb = openpyxl.load_workbook(path, read_only=False)
    first_sheet = wb.get_sheet_by_name('test case format')
    totalTestCases_counter = 0
    failTestCases_counter = 0
    passTestCases_counter = 0
    for i in range(1, first_sheet.max_row):
        if (first_sheet.cell(row=i, column=7).value) == 'fail':
            values[0] = values[0] + 1
        elif (first_sheet.cell(row=i, column=7).value) == 'pass':
            values[1] = values[1] + 1
    totalTestCases_counter = values[0] + values[1]
    print('Total teste fail: ', values[0])
    print('Total teste pass: ', values[1])
    print('Total teste: ', totalTestCases_counter)


def generateReport():
    wb = openpyxl.load_workbook(path, read_only=False)
    first_sheet = wb.get_sheet_by_name('test case format')
    tester = first_sheet['E3'].value
    try:
        reportSheet = wb.get_sheet_by_name("Report")
    except:
        wb.create_sheet("Report")
        reportSheet = wb.get_sheet_by_name("Report")

        reportSheet['A1'] = "TesterID: "
        reportSheet['B1'] = tester
        reportSheet["A2"] = "Failed test cases"
        reportSheet['B2'] = values[0]
        reportSheet['A3'] = "Passed test cases"
        reportSheet['B3'] = values[1]
        reportSheet['A4'] = "Total number of test cases"
        reportSheet['B4'] = values[0] + values[1]
        wb.save(path)
        createChart()


    excel = client.Dispatch('Excel.Application')
    sheets = excel.Workbooks.open(path)
    work_sheets = sheets.Worksheets[2]
    work_sheets.ExportAsFixedFormat(0, path_PDF)


def createChart():
    wb = openpyxl.load_workbook(path, read_only=False)
    sheet = wb['Report'] #reportSheet=wb.get_sheet_by_name('Report')
    pie = PieChart()

    labels = Reference(sheet, min_col=1, min_row = 2, max_row = 3)
    data = Reference(sheet, min_col=2, min_row = 2, max_row = 3)
    pie.add_data(data,titles_from_data=False)
    pie.set_categories(labels)
    pie.title = 'Test cases'

    pie.width = 14
    pie.height = 8

    sheet.add_chart(pie, "A6")
    wb.save(path)

def buttonPressed():
    compareValues()
    generateReport()

Button(root, text="Generate Report", command=buttonPressed).pack(pady=5)
root.mainloop()

